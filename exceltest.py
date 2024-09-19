from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
import os
from excel_letters import excel_letters
from dotenv import load_dotenv

# Load environment variables from the .env file
load_dotenv()

try:
	dirs = os.listdir("./output")

	trimprofiles = lambda ln: ln.strip("\n")

	profiles = filter(lambda file: True if "pro" in file else False, dirs)
	trimprofiles = lambda ln: ln.strip("\n")

	drops = []
	maxlength = 0


	entity = int(os.getenv('ENTITY'))

	# X Y starting

	rowIndex = 3
	colIndex = 2


	def sort_profiles(profiles):
		return profiles.sort()


	sorted_profiles = list(profiles)

	sheet_title = "sheet"  #input("Entity Name: ")


	def read_lines(path):
		global maxlength
		global drops
		f = open(path, "r")
		lines = f.readlines()
		trimmed = list(map(trimprofiles, lines))
		maxlength = len(lines) if len(lines) > maxlength else maxlength
		drops.append(trimmed)


	for fl in list(sorted_profiles):
		read_lines(f"./output/{fl}")

	# Create a workbook and select the active worksheet
	wb = Workbook()
	ws = wb.active
	ws.title = sheet_title
	f = open(f"./timedrops/CMH{entity}.txt", "r")
	timedrops = list(map(trimprofiles, f.readlines()))

	# Styling
	alignment = Alignment(horizontal="center", vertical="center")
	font = Font(bold=True, size=15)

	bold_border = Border(left=Side(style='thick'),
																						right=Side(style='thick'),
																						top=Side(style='thick'),
																						bottom=Side(style='thick'))

	# Create a fill pattern Black
	fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

	dropIndex = 0
	rdpcount = len(drops[0][0].strip().split('\t'))

	for drop in drops:
		ws.merge_cells(
						f"{excel_letters[colIndex-1]}{rowIndex}:{excel_letters[colIndex-1+rdpcount-1]}{rowIndex+1}"
		)
		ws[f"{excel_letters[colIndex-1]}{rowIndex}"] = timedrops[dropIndex]
		dropIndex += 1
		ws[f"{excel_letters[colIndex-1]}{rowIndex}"].font = font
		ws[f"{excel_letters[colIndex-1]}{rowIndex}"].alignment = alignment
		ws[f"{excel_letters[colIndex-1]}{rowIndex}"].border = bold_border

		for row in ws.iter_rows(min_row=3,
																										max_row=4,
																										min_col=colIndex,
																										max_col=colIndex + rdpcount - 1):
			for cell in row:
				cell.border = bold_border

		ws.column_dimensions[
						f"{excel_letters[colIndex-1 + rdpcount]}"].width = rdpcount

		for row in ws.iter_rows(min_row=1,
																										max_row=1000,
																										min_col=colIndex + rdpcount,
																										max_col=colIndex + rdpcount):
			for cell in row:
				cell.fill = fill

		rowIndexFilled = rowIndex + 2
		for ln in drop:
			item = ln.split('\t')
			for row in ws.iter_rows(min_row=rowIndexFilled,
																											max_row=rowIndexFilled,
																											min_col=colIndex,
																											max_col=colIndex + len(item) - 1):
				index = 0
				for cell in row:
					cell.value = int(item[index]) if item[index] != "" else ""
					cell.alignment = alignment
					index += 1
				rowIndexFilled += 1
		colIndex += rdpcount + 1
	wb.save("./output/sample.xlsx")
except Exception:
		print("Failed to Generate Excel File")